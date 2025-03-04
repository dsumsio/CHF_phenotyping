import ray
import re  # regular expression package
from tqdm import tqdm
from joblib import Parallel, delayed
import pandas as pd
import spacy
import csv
import json
from scipy.stats import mannwhitneyu
import numpy as np
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.drawing.image import Image

num_cpu = 15
samp_size = 1500
n_word_dist = 10

nlp = spacy.load('en_core_web_sm')
stopwords = nlp.Defaults.stop_words
 #Remove unmeaningful words by making use of common sense
nlp.Defaults.stop_words |= {"et","al","fig"}

 #Read notes_all file for patient notes
df_notes = pd.read_csv('notes_all.csv')
N = len(df_notes)

 #Sample random 10.000 notes from notes_all.csv
random = np.random.choice(N, samp_size, replace=False)
 #df_notes = df_notes.sample(n=10000)  #TODO delete. this is for testing only

text2 = df_notes['note_txt']
text1 = []
 #create list of random retrieved notes as dataset
for i in random:
     text1.append(text2[i])

 # this is code to do the data-driven approach to find keywords
 #Read the coresponding ICD/KW list

with open('KW_CHF.txt', 'r') as KW1:
    Kw = KW1.read()
with open('ICD_CHF.txt', 'r') as C_IS:
    ICD = C_IS.read()



#Start Ray
ray.init()

 ##Remove unmeaningful words and Lemmatize text
 # loop over each row in text1, and each row is an individual note
 #filtered_notes = []
 #for i in tqdm(range(len(text1))):

@ray.remote
def inner_loop1(note):
     #note = text1.iloc[i] # comment this in parallel
     doc2 = nlp(note)
     #Remove all unmeaningful words and create list of tuples of <token, stem> by lemmatization
     filtered_tokens = [(token.text, token.lemma_ ) for token in doc2 if not token.is_stop]
     #Create a list with only the meaningful stems of the words in text
     joined_sentence = [token[1].lower() for token in filtered_tokens]
     joined_sentence = " ".join(joined_sentence)
     return joined_sentence

 #with Parallel(n_jobs=num_cpu, backend ="threading", verbose=False) as par:
     #filtered_notes =par(delayed(inner_loop1)(text1.iloc[i]) for i in tqdm(range(len(text1))))

 # Run Ray Parallel
filtered_notes = ray.get([inner_loop1.remote(note) for note in text1])



 ##Lematize keywords
doc3 = nlp(Kw)
 #Remove all unmeaningful words and create list of tuples of <token, stem> by lemmatization
filtered_tokens2 = [(token.text, token.lemma_ ) for token in doc3 if not token.is_stop]
joined_sentence2 = [token[1].lower() for token in filtered_tokens2]
 #Create a list with only the meaningful stems of the words in text
filtered_notes2 = " ".join(joined_sentence2)

 ##Create n-grams
def ngrams(note, n):
     """
     The input of the function is a string (note)
     and the number of grams (n)
     The output is a list of vocabulary words that are unique words with length n
     """
     #note_string = ' '.join(note)
     doc = nlp(note)

     #creates a list for the result
     result = list()
     # create a list that contains no punctiation
     sentence = list()
     #parse through the document to pass all tokens that are words to sentence list

     for token in doc:
        #alphabetic character selection
        if token.is_alpha:
            sentence.append(token)

     #parse through the sentence while adding words in groups of two the result
     for word in range(len(sentence)-1):
         element = tuple(sentence[word:word+n])
         # think about 0:5 --> [0,1,2,3,4]
         result.append(element)

     return result

# because Jet does not have access to the dataset yet,
# therefore we use a small simulated dataset
"""notes = [
         'I feel sick want to go home. I\'m not good',
         'I have a headache and stiffness in neck',
         'my brain is bleeding',
         'I have a history of stroke',
         'I feel good',
         'I do not have stroke',
         'I have hypertension'
         ]"""

 #Cohort for expert labeling and model development: 1000 cases MGB, 1000 BIDMC
 #np.random.choice(N, 1000, replace=False)  

 #Use literature used to retrieve keywords to test code
 #split lines of text as seperate notes, because no access yet too patient data
 #notes = [x.strip() for x in filtered_notes.split('\n')]

 ## step 1: define positive and negative classes
 # we use (1) keywords for the `stroke_type` from literature and our common sense to define classes
 #        (2) ICD code as keywords for the `stroke_type`
 #initial_keywords = ['bleeding', 'sick', 'hemorrhage']

 #Add keywords/ICD-codes indivdually to list, put keep keywords consistent of multiple kw together
 #initial_keywords = Kw.split('\n')
initial_keywords = [x.strip() for x in filtered_notes2.split('\n')]
 ##Get list for Arjun's tool
 #for word in initial_keywords:
 #    key = ["/" + word + "/gmi,"]
 #    key2 = " ".join(key)
 #    print(key2)

initial_icd=ICD.split('\n')
res_initial = initial_icd[:-1]
 ##Get list for Arjun's tool
 #for word in initial_icd:
 #    key = ["/" + word + "/gmi,"]
 #    key2 = " ".join(key)
 #    print(key2)

initial_kw_icd_pattern = re.compile('(?:' + '|'.join(initial_keywords+res_initial) + ')', re.I)
 #initial_kw_icd_pattern = re.compile('(?:' + '|'.join(initial_keywords+initial_icd) + ')', re.I)

 # for each note, we decide if it contains any of the initial keyword or ICD code, then it's positive class
 # otherwise it's negative class
notes_class = []
for note in tqdm(filtered_notes):
     #note_string = ' '.join(note)
     if re.search(initial_kw_icd_pattern, note):
         notes_class.append(1)
     else:
         notes_class.append(0)
cls = notes_class
     # check if it contains any initial keywords or ICDs
     # the below is a list of booleans
     # for example, it can be [True, False] --> bleeding is in the note and sick is not
     #cls = int(any([k in note for k in initial_keywords]))  # positive -> 1 and negative -> 0
     # this is the fancy regular expression way to do the same thing as above

 # up to now, we have a list `notes_class` of integers 0/1, indicating the class of each note

 ## step 2: use "inverse of BoW" to find words that are most different across the positive and negative classes
 # we need a p-value by comparing the distributions of the word count frequencies

 # first, we need the vocabulory, which is a list of unique (n-) words
 # for example, "how are you". if n = 1 ---> unique_words = ['how', 'are', 'you']
 # for example, "how are you". if n = 2 ---> unique_words = ['how are', 'are you']
 # for example, "how are you". if n = 3 ---> unique_words = ['how are you']
n = 2
 #empty set to collect the unique words
 #vocab = set()

 # notes is a list of strings
 #for note in tqdm(filtered_notes):

@ray.remote
def inner_loop2(note):
     # note is a (long) string
     # n is the number of n-gram
     # result is a list of tuples, each tuple has n elements, each element is a Token
     result = ngrams(note, n)
     token_vocab = []
     for token in result:  # `token`` is a tuple of n Token's
             t = tuple([str(item) for item in token]) # each `item` is a Token, use str(item) to convert it into string
         # we want to add a tuple of string's, so that one tuple of string's is comparable to another tuple of string's
         # Check if string length > 1, to exclude single letters to vocab
             if len(t[0]) > 1:
                 token_vocab.append(t)
     return token_vocab

vocab = ray.get([inner_loop2.remote(note) for note in filtered_notes])                
#vocab = [inner_loop2(note) for note in filtered_notes]

 #with Parallel(n_jobs=num_cpu, backend ="threading", verbose=False) as par:
 #    vocab = par(delayed(inner_loop2)(note) for note in tqdm(filtered_notes))


vocab = list(vocab)
     # if n=1, then result = ['I', 'have', 'hypertension', 'sick', ...]
     # if n=2, then result = ['I have', 'have hypertension', ...]
     # however, since now `token`` is a tuple where its elements have type=spacy.token
     # previously, using token.text worked because `token` is not a tuple, but it is spacy.token
     # therefore we can use .text function to get its text (type=str)
    # for token in result:
     #    if token not in result:
    #         vocab.add(token)

 #unique_words = []
 #for note in notes:
 #   words = note.split('\s')
     # words = ['how', 'are', 'you']

     # how are you
     # n=1:
     #  1   2   3
     # n=2:
     #  1   1
     #      2   2
     # n=3:
     #  1   1   1

     # in case len(words)=3 and n=1, this is range(3) which is [0,1,2]
     # in case len(words)=3 and n=2, this is range(2) which is [0,1]
#for i in range(len(words)-n+1):
#unique_words = list(set(unique_words))  # make them unique

#get unique words
# Flatten the nested list into a single list
flattened_list = [word for sublist in vocab for word in sublist]

# Convert the flattened list to a set to remove duplicates
unique_words_set = set(flattened_list)

# Convert the set back to a list
unique_vocab = list(unique_words_set)



## then, for each word in the vocabulory, get it's count distribution in the positive class
@ray.remote
def inner_loop3(element):
    n_gram = ''.join([token + ' ' for token in element]).rstrip()
    pos_cnt = []
    neg_cnt = []

    for note, cls in zip(filtered_notes, notes_class):
         ct = note.count(n_gram)
         if cls==1:
             pos_cnt.append(ct)
         else:
             neg_cnt.append(ct)

     # get p-value if all cls == 1 or cls == 0
    if not neg_cnt:
         neg_cnt = [0]
    if not pos_cnt:
         pos_cnt = [0]

    pval = mannwhitneyu(pos_cnt, neg_cnt).pvalue
    pos_frequency = sum(pos_cnt)/sum(notes_class)
    neg_frequency = sum(neg_cnt)/(samp_size - sum(notes_class))
    return pval, pos_cnt, neg_cnt, pos_frequency, neg_frequency, n_gram

results = ray.get([inner_loop3.remote(note) for note in unique_vocab])

"""
resutlts = results_pre['pval']
sorted_list = {
    key: [x for _, x in sorted(zip(sorting_list, sublist))] 
    for key, sublist in my_list.items()
}
"""

pvals = [result[0] for result in results]
pos_dist = [result[1] for result in results]
neg_dist = [result[2] for result in results]
pos_freq = [result[3] for result in results]
neg_freq = [result[4] for result in results]
ngrm = [result[5] for result in results]

# save results in a pandas dataframe
df = pd.DataFrame(data={
     'word':unique_vocab,
     'pval':pvals,
     'Positive frequency':pos_freq,
     'Negative frequency':neg_freq
     })
 # rank `df` according to pvalue in ascending order
df = df.sort_values('pval', ascending=True, ignore_index=True)

## Create a figure and set size
fig = plt.figure(figsize=(18,12))

#set x-step size, length of total ngrams
xloc = np.arange(len(ngrm[:n_word_dist]))
#Define width of each bar 
bar_width = 0.25
#position middle of the bars
xloc1 = np.arange(len(ngrm[:n_word_dist])) - bar_width/2
xloc2 = np.arange(len(ngrm[:n_word_dist])) + bar_width/2
yloc1 = df['Positive frequency'][:n_word_dist]
yloc2 = df['Negative frequency'][:n_word_dist]
#Plot pos/neg frequencies 
plt.bar(xloc1, yloc1, width=bar_width)
plt.bar(xloc2, yloc2, width=bar_width)
#define label location
plt.xticks(ticks=xloc, labels=ngrm[:n_word_dist])
plt.xlabel('Ngram')
plt.ylabel('Frequency')
plt.legend(['positive frequency', 'negative frequency'])
plt.title('The positive and negative frequency distribution of 10 ngrams over all notes')

plt.show()

# Stop Parallelization
ray.shutdown()

 # #Only retrieve first 200 lines
 # df2 = df.head(200)
 # df2.to_excel(f'inverse_BoW_result_{n}gramDATA_SDH_KW_200.xlsx', index=False)
 # result = pd.read_excel(f'inverse_BoW_result_{n}gramDATA_SDH_KW_200.xlsx')
 # print(df2)
df.to_excel('DDM_withplot.xlsx', sheet_name='Data', index=False)

# Save the plot, then load it onto the excel sheet
plt.savefig('histogram.png')
plot_image = Image('histogram.png')
workbook = load_workbook('DDM_withplot.xlsx')
plot_sheet = workbook.create_sheet(title='Plot')
plot_sheet.add_image(plot_image, 'A1')
workbook.save('DDM_withplot.xlsx')

#result = pd.read_excel(f'inverse_BoW_result_{n}gramDATA_IS_KW_try.xlsx')
 


 
print(df)